"""
Parse CSV and XLSX bank export files.
Attempts to auto-detect columns; falls back to manual mapping.
"""
import csv
import hashlib
import re
import pandas as pd
from io import BytesIO, StringIO


COMMON_DATE_COLS = ["date", "transaction date", "posted date", "trans date", "trans. date", "posting date"]
COMMON_DESC_COLS = ["description", "payee", "memo", "name", "transaction description", "details"]
COMMON_AMOUNT_COLS = ["amount", "transaction amount", "debit/credit"]
COMMON_DEBIT_COLS = ["debit", "withdrawal", "withdrawals", "charges"]
COMMON_CREDIT_COLS = ["credit", "deposit", "deposits", "payments"]


def _normalize(s: str) -> str:
    return str(s).lower().strip()


def _find_col(df_cols: list[str], candidates: list[str]) -> str | None:
    norm_cols = {_normalize(c): c for c in df_cols}
    for cand in candidates:
        if _normalize(cand) in norm_cols:
            return norm_cols[_normalize(cand)]
    return None


def _detect_header_row(raw) -> int | None:
    """
    Index of the row that looks like the column-header row — i.e. the first row
    (within the first 25) that contains a recognizable Date column. Lets us skip
    the title/company/report-name rows many bank & QuickBooks exports put on top.
    """
    for i in range(min(25, len(raw))):
        cells = [_normalize(c) for c in raw.iloc[i].tolist()]
        cells = [c for c in cells if c and c != "nan"]
        if cells and any(c in COMMON_DATE_COLS for c in cells):
            return i
    return None


def _parse_amount(val) -> float | None:
    if pd.isna(val) or str(val).strip().lower() in ("", "-", "n/a", "nan", "none"):
        return None
    s = re.sub(r"[,$\s]", "", str(val))
    s = s.replace("(", "-").replace(")", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_spreadsheet(file_bytes: bytes, filename: str, account_id: str,
                      col_mapping: dict = None) -> tuple[list[dict], list[str]]:
    """
    Parse CSV or XLSX.
    Returns (transactions, warnings).
    col_mapping can override auto-detection: {"date": "Date", "payee": "Payee", ...}
    """
    warnings = []

    # Read the file as a raw grid (no header) so we can locate the real header
    # row even when title/preamble rows sit on top (common in bank & QB exports).
    if filename.lower().endswith((".xlsx", ".xls")):
        raw = pd.read_excel(BytesIO(file_bytes), dtype=str, header=None)
    else:
        # Use the csv module: pandas' header=None keys the column count off the
        # first line, so a short title row on top would make it drop the wider
        # header/data rows. Reading rows ourselves and padding avoids that.
        text = file_bytes.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(StringIO(text)))
        width = max((len(r) for r in rows), default=0)
        rows = [r + [""] * (width - len(r)) for r in rows]
        raw = pd.DataFrame(rows, dtype=str)

    if raw.empty:
        return [], ["The file appears to be empty."]

    header_idx = _detect_header_row(raw)
    if header_idx is None:
        header_idx = 0  # fall back to treating the first row as the header

    header_cells = [_normalize(c) for c in raw.iloc[header_idx].tolist()]

    # QuickBooks "Transaction Detail by Account" has a 'Split' column and groups
    # rows under account section-headers rather than a per-row account column.
    # This single-account importer can't represent that — send the user to the
    # wizard that's built for it instead of silently importing nothing.
    if "split" in header_cells:
        raise ValueError(
            "This looks like a QuickBooks 'Transaction Detail by Account' "
            "export (it has a 'Split' column and groups rows by account).\n\n"
            "Import it from Settings → Open Import Wizard → Step 2 "
            "(Transactions), which is built for that multi-account format.")

    # Re-frame the table using the detected header row.
    df = raw.iloc[header_idx + 1:].copy()
    df.columns = [str(c).strip() for c in raw.iloc[header_idx].tolist()]
    df = df.reset_index(drop=True)

    if col_mapping:
        date_col = col_mapping.get("date")
        desc_col = col_mapping.get("payee")
        amount_col = col_mapping.get("amount")
        debit_col = col_mapping.get("debit")
        credit_col = col_mapping.get("credit")
    else:
        date_col = _find_col(df.columns, COMMON_DATE_COLS)
        desc_col = _find_col(df.columns, COMMON_DESC_COLS)
        amount_col = _find_col(df.columns, COMMON_AMOUNT_COLS)
        debit_col = _find_col(df.columns, COMMON_DEBIT_COLS)
        credit_col = _find_col(df.columns, COMMON_CREDIT_COLS)

    if not date_col:
        warnings.append("Could not detect a Date column.")
    if not desc_col:
        warnings.append("Could not detect a Description/Payee column.")
    if not amount_col and not (debit_col or credit_col):
        warnings.append("Could not detect an Amount column.")

    results = []
    for idx, row in df.iterrows():
        date_val = str(row.get(date_col, "") if date_col else "").strip()
        if not date_val or date_val.lower() in ("nan", ""):
            continue

        # Normalize date
        try:
            date_str = pd.to_datetime(date_val, dayfirst=False).strftime("%Y-%m-%d")
        except Exception:
            warnings.append(f"Row {header_idx+idx+2}: unrecognized date '{date_val}', skipped.")
            continue

        payee = str(row.get(desc_col, "") if desc_col else "").strip()

        # Resolve amount
        amount = None
        if amount_col:
            amount = _parse_amount(row.get(amount_col, ""))
        elif debit_col or credit_col:
            debit = _parse_amount(row.get(debit_col, "") if debit_col else None) or 0.0
            credit = _parse_amount(row.get(credit_col, "") if credit_col else None) or 0.0
            # Debits are outflows (negative), credits are inflows (positive)
            amount = credit - debit if (credit or debit) else None

        if amount is None:
            warnings.append(f"Row {header_idx+idx+2}: no amount found, skipped.")
            continue

        raw = f"{account_id}|{date_str}|{amount}|{payee}"
        import_hash = hashlib.md5(raw.encode()).hexdigest()[:16]

        results.append({
            "date": date_str,
            "account_id": account_id,
            "payee": payee,
            "memo": "",
            "amount": str(round(amount, 2)),
            "category_id": "",
            "class_id": "",
            "is_transfer": "0",
            "transfer_pair_id": "",
            "reconciled": "0",
            "reconcile_id": "",
            "notes": "",
            "import_hash": import_hash,
        })

    return results, warnings


def parse_qb_multi_account_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse a QuickBooks 'Transaction Detail by Account' Excel export.

    QB formats this report with account names as section header rows between
    groups of transactions — there is no 'Account' column on each row.
    The file may contain multiple sheets; this function auto-detects the one
    with transaction data by looking for a row that contains 'Date'.
    Returns flat transaction dicts each with an '_account_name' field.
    """
    warnings: list[str] = []
    transactions: list[dict] = []

    # ── Auto-detect which sheet has the data ──────────────────────────────────
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    target_sheet = 0
    for idx, sname in enumerate(wb.sheetnames):
        ws = wb[sname]
        for row in ws.iter_rows(max_row=10, values_only=True):
            if any(str(c or "").strip().lower() == "date" for c in row):
                target_sheet = idx
                break
        else:
            continue
        break
    wb.close()

    df_raw = pd.read_excel(BytesIO(file_bytes), dtype=str, header=None,
                           sheet_name=target_sheet)

    # ── Find the column-header row ────────────────────────────────────────────
    header_idx = None
    date_col = name_col = memo_col = split_col = None
    debit_col = credit_col = amount_col = None

    for i, row in df_raw.iterrows():
        cells_lower = [str(c).strip().lower() for c in row]
        if "date" in cells_lower:
            header_idx = i
            for j, cell in enumerate(cells_lower):
                if cell == "date"   and date_col   is None: date_col   = j
                if cell in ("name", "payee", "description") and name_col   is None: name_col   = j
                if cell == "memo"                           and memo_col   is None: memo_col   = j
                if cell == "split"                         and split_col  is None: split_col  = j
                if cell in ("debit", "withdrawal")         and debit_col  is None: debit_col  = j
                if cell in ("credit", "deposit")           and credit_col is None: credit_col = j
                if cell == "amount"                        and amount_col is None: amount_col = j
            break

    if header_idx is None or date_col is None:
        warnings.append(
            "Could not find a header row with a 'Date' column. "
            "Make sure you exported 'Transaction Detail by Account' to Excel.")
        return [], warnings

    # ── Walk data rows ────────────────────────────────────────────────────────
    current_account = ""

    for i, row in df_raw.iterrows():
        if i <= header_idx:
            continue

        cells = [str(c).strip() for c in row]

        # Skip fully empty rows
        if all(c.lower() in ("", "nan", "none") for c in cells):
            continue

        # First non-empty cell — account headers can be in any column
        first_val = ""
        for c in cells:
            if c.lower() not in ("", "nan", "none"):
                first_val = c
                break

        date_val = cells[date_col] if date_col < len(cells) else ""

        # Skip total / subtotal / balance summary rows
        if first_val.lower().startswith("total") or first_val.lower() == "balance":
            continue

        # Try parsing the date column
        date_str = None
        if date_val and date_val.lower() not in ("nan", "none", ""):
            try:
                date_str = pd.to_datetime(date_val, dayfirst=False).strftime("%Y-%m-%d")
            except Exception:
                pass

        if date_str:
            # ── Transaction row ───────────────────────────────────────────────
            if not current_account:
                warnings.append(
                    f"Transaction on {date_str} has no account context — skipped.")
                continue

            payee = ""
            if name_col is not None and name_col < len(cells):
                v = cells[name_col]
                payee = "" if v.lower() in ("nan", "none") else v

            memo = ""
            if memo_col is not None and memo_col < len(cells):
                v = cells[memo_col]
                memo = "" if v.lower() in ("nan", "none") else v

            split = ""
            if split_col is not None and split_col < len(cells):
                v = cells[split_col]
                split = "" if v.lower() in ("nan", "none") else v

            amount = None
            if amount_col is not None and amount_col < len(cells):
                amount = _parse_amount(cells[amount_col])
            if amount is None and (debit_col is not None or credit_col is not None):
                debit  = (_parse_amount(cells[debit_col])  if debit_col  is not None and debit_col  < len(cells) else None) or 0.0
                credit = (_parse_amount(cells[credit_col]) if credit_col is not None and credit_col < len(cells) else None) or 0.0
                if credit or debit:
                    amount = credit - debit

            if amount is None:
                warnings.append(f"Row {i + 2}: no amount found — skipped.")
                continue

            raw = f"{current_account}|{date_str}|{amount}|{payee}"
            import_hash = hashlib.md5(raw.encode()).hexdigest()[:16]

            transactions.append({
                "_account_name":    current_account,
                "_split":           split,   # QB offsetting account → used as category
                "date":             date_str,
                "payee":            payee,
                "memo":             memo,
                "amount":           str(round(amount, 2)),
                "category_id":      "",
                "class_id":         "",
                "is_transfer":      "0",
                "transfer_pair_id": "",
                "reconciled":       "0",
                "reconcile_id":     "",
                "notes":            "",
                "import_hash":      import_hash,
            })
        else:
            # ── Possible account section header ───────────────────────────────
            # Date column is empty; if there's any non-empty non-total cell it's
            # an account name.
            if first_val and first_val.lower() not in ("nan", "none"):
                current_account = first_val

    return transactions, warnings


def get_columns(file_bytes: bytes, filename: str) -> list[str]:
    """Return list of column names for column-mapping UI."""
    if filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls"):
        df = pd.read_excel(BytesIO(file_bytes), dtype=str, nrows=1)
    else:
        df = pd.read_csv(BytesIO(file_bytes), dtype=str, nrows=1, on_bad_lines="skip")
    return [str(c).strip() for c in df.columns]
